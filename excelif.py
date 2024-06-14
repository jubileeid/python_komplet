from openpyxl import Workbook, load_workbook

wb = Workbook()
ws = wb.active
ws.title = "Data Baru"

ws["A1"] = "L"
ws["B1"] = "P"
ws["C1"] = """=IF(A1="L","Laki-Laki","Perempuan")"""

wb.save("if.xlsx")
