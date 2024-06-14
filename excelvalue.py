from openpyxl import workbook, load_workbook

wb = load_workbook('datasales.xlsx')
ws = wb.active

nomor = ws["A3"].value
sales = ws["B3"].value

print("Sales dengan nomor {0} adalah {1}" .format(nomor,sales))
